import os
from flask import Flask, render_template, request, redirect, url_for, send_file, session
import sqlite3
import openpyxl
from datetime import datetime
import hashlib
from pathlib import Path

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'ваш_секретный_ключ')


# Хешируем пароль
def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()


# Правильный пароль (хеш от "admin123")
CORRECT_PASSWORD_HASH = hash_password("admin123")


# Проверка авторизации
def check_admin():
    return session.get('is_admin') == True


# Путь к базе данных
def get_db_path():
    # На Render используем /tmp для SQLite
    if 'RENDER' in os.environ:
        return '/tmp/database.db'
    else:
        return 'database.db'


# Создаем базу данных
def init_db():
    db_path = get_db_path()
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS анкеты (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            фио TEXT,
            дата_рождения TEXT,
            возраст INTEGER,
            пол TEXT,
            телефон TEXT,
            email TEXT,
            гражданство TEXT,
            образование TEXT,
            опыт_переписей TEXT,
            подробности_опыта TEXT,
            желаемая_должность TEXT,
            район_работы TEXT,
            знакомство_с_проектами TEXT,
            мотивация TEXT,
            психологический_тест TEXT,
            расшифровка_теста TEXT,
            дата_заполнения TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()


# Инициализация при старте
init_db()


@app.route('/')
def анкета():
    return render_template('form.html')


@app.route('/сохранить', methods=['POST'])
def сохранить():
    # Собираем данные
    данные = {
        'фио': request.form['фио'],
        'дата_рождения': request.form['дата_рождения'],
        'возраст': request.form['возраст'],
        'пол': request.form['пол'],
        'телефон': request.form['телефон'],
        'email': request.form['email'],
        'гражданство': request.form['гражданство'],
        'образование': request.form['образование'],
        'опыт_переписей': request.form.get('опыт_переписей', 'Нет'),
        'подробности_опыта': request.form.get('подробности_опыта', ''),
        'желаемая_должность': request.form['желаемая_должность'],
        'район_работы': request.form['район_работы'],
        'знакомство_с_проектами': request.form['знакомство_с_проектами'],
        'мотивация': request.form['мотивация']
    }

    # Психологический тест
    выбранные_варианты = request.form.getlist('психологический_тест')
    психологический_тест = ', '.join(выбранные_варианты)

    # Расшифровка
    расшифровки = {
        '1': 'Ориентирован на развитие и коллективное интеллектуальное обогащение',
        '2': 'Потребность в видимом прогрессе, важно видеть результат команды',
        '3': 'Стремление к гармонии и снижению эмоциональной усталости',
        '4': 'Интроверт/амбиверт, эффективен при минимальных отвлечениях',
        '5': 'Эмоциональные ценности, работа - часть жизни с поддержкой близких',
        '6': 'Организованность и ориентация на результат',
        '7': 'Социальная направленность, потребность в неформальном взаимодействии',
        '8': 'Ценность личного пространства, концентрация, минимальный шум',
        '9': 'Аналитический и решительный подход к задачам'
    }

    расшифровка_теста = ' | '.join([f"Вариант {номер}: {расшифровки.get(номер, 'Неизвестно')}"
                                    for номер in выбранные_варианты])

    # Сохраняем
    db_path = get_db_path()
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute('''
        INSERT INTO анкеты 
        (фио, дата_рождения, возраст, пол, телефон, email, гражданство, 
         образование, опыт_переписей, подробности_опыта, желаемая_должность,
         район_работы, знакомство_с_проектами, мотивация,
         психологический_тест, расшифровка_теста)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        данные['фио'], данные['дата_рождения'], данные['возраст'],
        данные['пол'], данные['телефон'], данные['email'],
        данные['гражданство'], данные['образование'], данные['опыт_переписей'],
        данные['подробности_опыта'], данные['желаемая_должность'],
        данные['район_работы'], данные['знакомство_с_проектами'],
        данные['мотивация'], психологический_тест, расшифровка_теста
    ))
    conn.commit()
    conn.close()

    return redirect('/спасибо')


@app.route('/спасибо')
def спасибо():
    return render_template('thank_you.html')


# Админка
@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        password = request.form.get('password', '')
        if hash_password(password) == CORRECT_PASSWORD_HASH:
            session['is_admin'] = True
            return redirect('/admin')
        else:
            return render_template('admin_login.html', error='Неверный пароль')

    return render_template('admin_login.html')


@app.route('/admin/logout')
def admin_logout():
    session.pop('is_admin', None)
    return redirect('/')


@app.route('/admin')
def admin_panel():
    if not check_admin():
        return redirect('/admin/login')

    db_path = get_db_path()
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute('SELECT * FROM анкеты ORDER BY дата_заполнения DESC')
    все_анкеты = c.fetchall()
    conn.close()

    return render_template('admin.html', анкеты=все_анкеты, count=len(все_анкеты))


@app.route('/admin/download')
def download_excel():
    if not check_admin():
        return redirect('/admin/login')

    # Создаем Excel
    db_path = get_db_path()
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute('SELECT * FROM анкеты')
    данные = c.fetchall()

    c.execute('PRAGMA table_info(анкеты)')
    колонки = [колонка[1] for колонка in c.fetchall()]
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Анкеты"

    # Заголовки
    for номер_колонки, название in enumerate(колонки, 1):
        ws.cell(row=1, column=номер_колонки, value=название)

    # Данные
    for номер_строки, строка in enumerate(данные, 2):
        for номер_колонки, значение in enumerate(строка, 1):
            ws.cell(row=номер_строки, column=номер_колонки, value=значение)

    # Сохраняем
    имя_файла = f'анкеты_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(имя_файла)

    return send_file(имя_файла, as_attachment=True)


if __name__ == '__main__':
    # На Render используется gunicorn, этот блок только для локального запуска
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)